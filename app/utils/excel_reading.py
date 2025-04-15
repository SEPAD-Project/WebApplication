import openpyxl
from datetime import time
from flask_login import current_user

from app.utils.generate_class_code import generate_class_code


def add_students(
    path_to_xlsx, sheet_name, name_letter, family_letter, nc_letter, class_letter, pass_letter, available_classes, registered_national_codes
):
    """
    Adds students from an Excel file to the database, verifying format and preventing duplicates.

    Args:
        path_to_xlsx (str): Path to the Excel file containing student data.
        sheet_name (str): Name of the sheet in the Excel file.
        name_letter (str): Column letter for student names.
        family_letter (str): Column letter for student family names.
        nc_letter (str): Column letter for student national codes.
        class_letter (str): Column letter for class codes.
        pass_letter (str): Column letter for student passwords.
        available_classes (list): List of valid class codes.
        registered_national_codes (list): List of national codes already registered in the system.

    Returns:
        list or str: A list of student dictionaries if the data is valid, or an error message/list of problems.
    """

    # Load the workbook and select the specified sheet
    workbook = openpyxl.load_workbook(path_to_xlsx)
    try:
        sheet = workbook[sheet_name]
    except KeyError:
        return 'sheet_not_found'  # Return error if the sheet is not found

    # Convert column letters to zero-based indices
    try:
        name_index = openpyxl.utils.column_index_from_string(name_letter) - 1
        family_index = openpyxl.utils.column_index_from_string(
            family_letter) - 1
        nc_index = openpyxl.utils.column_index_from_string(nc_letter) - 1
        class_index = openpyxl.utils.column_index_from_string(class_letter) - 1
        pass_index = openpyxl.utils.column_index_from_string(pass_letter) - 1
    except ValueError:
        return 'bad_column_letter'  # Return error if invalid column letters are provided

    # Validate the data format and check for issues
    problems = []
    nc_list = []

    # Validate names and family names
    for cell in sheet[name_letter]:
        if cell.row == 1:
            continue
        if not isinstance(cell.value, str):
            problems.append(['bad_format', cell.row, cell.column_letter])

    for cell in sheet[family_letter]:
        if cell.row == 1:
            continue
        if not isinstance(cell.value, str):
            problems.append(['bad_format', cell.row, cell.column_letter])

    # Validate national codes
    for cell in sheet[nc_letter]:
        if cell.row == 1:
            continue
        if not isinstance(cell.value, int):
            problems.append(['bad_format', cell.row, cell.column_letter])
        elif str(cell.value) in registered_national_codes or str(cell.value) in nc_list:
            problems.append(['duplicated_nc', cell.row, cell.column_letter])
        else:
            nc_list.append(str(cell.value))

    # Validate class codes
    for cell in sheet[class_letter]:
        if cell.row == 1:
            continue
        if not isinstance(cell.value, (str, int)):
            problems.append(['bad_format', cell.row, cell.column_letter])
        elif str(cell.value) not in available_classes:
            problems.append(['unknown_class', cell.row, cell.column_letter])

    # Validate passwords
    for cell in sheet[pass_letter]:
        if cell.row == 1:
            continue
        if not isinstance(cell.value, int):
            problems.append(['bad_format', cell.row, cell.column_letter])

    # If any problems are detected, return them
    if problems:
        return problems

    # Extract valid student data
    students = []
    school_code = current_user.school_code

    for row in sheet.iter_rows(values_only=True, min_row=2):
        name = row[name_index]
        family = row[family_index]
        nc = row[nc_index]
        class_ = generate_class_code(school_code, str(row[class_index]))
        password = row[pass_index]

        students.append({
            'name': name,
            'family': family,
            'national_code': nc,
            'class': class_,
            'password': password
        })

    return students


def add_classes(path_to_xlsx, sheet_name, name_letter, available_classes):
    """
    Adds class data from an Excel file, checking for duplicates and format errors.

    Args:
        path_to_xlsx (str): Path to the Excel file containing class data.
        sheet_name (str): Name of the sheet in the Excel file.
        name_letter (str): Column letter for class names.
        available_classes (list): List of class codes already available in the system.

    Returns:
        list or str: A list of class dictionaries if the data is valid, or an error message/list of problems.
    """

    # Load the workbook and select the specified sheet
    workbook = openpyxl.load_workbook(path_to_xlsx)
    try:
        sheet = workbook[sheet_name]
    except KeyError:
        return 'sheet_not_found'  # Return error if the sheet is not found

    # Convert column letter to zero-based index
    try:
        name_index = openpyxl.utils.column_index_from_string(name_letter) - 1
    except ValueError:
        return 'bad_column_letter'  # Return error if invalid column letter is provided

    # Validate the data format and check for issues
    problems = []
    names_list = []

    for cell in sheet[name_letter]:
        if cell.row == 1:
            continue
        if not isinstance(cell.value, (str, int)):
            problems.append(['bad_format', cell.row, cell.column_letter])
        elif str(cell.value) in available_classes or str(cell.value) in names_list:
            problems.append(['duplicated_name', cell.row, cell.column_letter])
        else:
            names_list.append(str(cell.value))

    # If any problems are detected, return them
    if problems:
        return problems

    # Extract valid class data
    classes = []
    school_code = current_user.school_code

    for row in sheet.iter_rows(values_only=True, min_row=2):
        name = row[name_index]
        code = generate_class_code(school_code, str(name))
        classes.append({'name': str(name), 'code': code})

    return classes


def schedule_extraction(path_to_xlsx, sheet_name):
    """
    Extracts a schedule from an Excel file.

    The schedule is represented as a dictionary where each key corresponds to a row identifier (weekday),
    and the value is another dictionary mapping column headers (time range) to cell values (teacher national code).

    Args:
        path_to_xlsx (str): Path to the Excel file containing the schedule.
        sheet_name (str): Name of the sheet in the Excel file.

    Returns:
        dict or str: A dictionary representing the schedule, or an error message if the sheet is not found.
    """

    # Load the workbook and select the specified sheet
    workbook = openpyxl.load_workbook(path_to_xlsx)

    try:
        sheet = workbook[sheet_name]
    except KeyError:
        return 'sheet_not_found'  # Return error if the sheet is not found

    # Extract the schedule data
    schedule = {}

    for row_index, row in enumerate(sheet.iter_rows(values_only=True, min_row=2)):
        schedule[str(row[0])] = {}
        for column in sheet.iter_cols(values_only=True, min_col=2):
            schedule[str(row[0])][str(column[0])] = column[row_index + 1]

    return schedule
