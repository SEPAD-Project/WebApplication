# Third-party Imports
import openpyxl
from flask_login import current_user

# Internal Imports
from source.utils.generate_class_code import generate_class_code


def add_students(
    path_to_xlsx, sheet_name, name_letter, family_letter,
    nc_letter, class_letter, pass_letter, phone_letter,
    available_classes, registered_national_codes, registered_phone_numbers
):
    """
    Parse and validate student data from an Excel file.

    Args:
        path_to_xlsx (str): Path to the Excel file.
        sheet_name (str): Sheet name within the Excel file.
        name_letter (str): Column letter for student names.
        family_letter (str): Column letter for family names.
        nc_letter (str): Column letter for national codes.
        class_letter (str): Column letter for class names.
        pass_letter (str): Column letter for passwords.
        available_classes (list): Valid class names to match against.
        registered_national_codes (list): Existing national codes in DB.

    Returns:
        list[dict] or list[list] or str: Parsed student list, or errors, or a string error code.
    """
    try:
        workbook = openpyxl.load_workbook(path_to_xlsx)
        sheet = workbook[sheet_name]
    except KeyError:
        return 'sheet_not_found'

    # Convert column letters to zero-based indices
    try:
        name_index = openpyxl.utils.column_index_from_string(name_letter) - 1
        family_index = openpyxl.utils.column_index_from_string(
            family_letter) - 1
        nc_index = openpyxl.utils.column_index_from_string(nc_letter) - 1
        class_index = openpyxl.utils.column_index_from_string(class_letter) - 1
        pass_index = openpyxl.utils.column_index_from_string(pass_letter) - 1
        phone_index = openpyxl.utils.column_index_from_string(phone_letter) - 1
    except ValueError:
        return 'bad_column_letter'

    problems = []
    nc_list = []
    phone_list = []

    for row in sheet.iter_rows(min_row=2):
        # Fetch raw cell values
        name = row[name_index].value
        family = row[family_index].value
        nc = row[nc_index].value
        class_raw = row[class_index].value
        password = row[pass_index].value
        phone = row[phone_index].value

        row_number = row[0].row

        # Validate name and family
        if not isinstance(name, str):
            problems.append(['bad_format', row_number, name_letter])
        if not isinstance(family, str):
            problems.append(['bad_format', row_number, family_letter])

        # Validate national code
        if not isinstance(nc, int):
            problems.append(['bad_format', row_number, nc_letter])
        elif str(nc) in registered_national_codes or str(nc) in nc_list:
            problems.append(['duplicated_nc', row_number, nc_letter])
        else:
            nc_list.append(str(nc))

        # Validate class
        if not isinstance(class_raw, (str, int)):
            problems.append(['bad_format', row_number, class_letter])
        elif str(class_raw) not in available_classes:
            problems.append(['unknown_class', row_number, class_letter])

        # Validate password
        if not isinstance(password, int):
            problems.append(['bad_format', row_number, pass_letter])

        # Validate phone number
        if not isinstance(phone, int):
            problems.append(['bad_format', row_number, phone_letter])
        elif str(phone) in registered_phone_numbers or str(phone) in phone_list:
            problems.append(['duplicated_nc', row_number, phone_list])
        else:
            phone_list.append(str(phone))

    if problems:
        return problems

    # Build clean data
    students = []
    school_code = current_user.school_code

    for row in sheet.iter_rows(values_only=True, min_row=2):
        students.append({
            'name': row[name_index],
            'family': row[family_index],
            'national_code': row[nc_index],
            'class': generate_class_code(school_code, str(row[class_index])),
            'password': row[pass_index],
            'phone_number': row[phone_index]
        })

    return students


def add_classes(path_to_xlsx, sheet_name, name_letter, available_classes):
    """
    Parse and validate class names from Excel input.

    Args:
        path_to_xlsx (str): Path to the Excel file.
        sheet_name (str): Sheet to read.
        name_letter (str): Column letter for class names.
        available_classes (list): Existing class names to avoid duplication.

    Returns:
        list[dict] or list[list] or str: Parsed class list, or errors, or a string error code.
    """
    try:
        workbook = openpyxl.load_workbook(path_to_xlsx)
        sheet = workbook[sheet_name]
    except KeyError:
        return 'sheet_not_found'

    try:
        name_index = openpyxl.utils.column_index_from_string(name_letter) - 1
    except ValueError:
        return 'bad_column_letter'

    problems = []
    names_list = []

    for row in sheet.iter_rows(min_row=2):
        name = row[name_index].value
        row_number = row[0].row

        # Check format and duplication
        if not isinstance(name, (str, int)):
            problems.append(['bad_format', row_number, name_letter])
        elif str(name) in available_classes or str(name) in names_list:
            problems.append(['duplicated_name', row_number, name_letter])
        else:
            names_list.append(str(name))

    if problems:
        return problems

    # Build valid class data
    school_code = current_user.school_code
    classes = []

    for row in sheet.iter_rows(values_only=True, min_row=2):
        name = str(row[name_index])
        code = generate_class_code(school_code, name)
        classes.append({'name': name, 'code': code})

    return classes


def schedule_extraction(path_to_xlsx, sheet_name):
    """
    Extract structured schedule data from an Excel file.

    Args:
        path_to_xlsx (str): Excel file path.
        sheet_name (str): Sheet name to read from.

    Returns:
        dict or str: Schedule dictionary or 'sheet_not_found' error.
    """
    try:
        workbook = openpyxl.load_workbook(path_to_xlsx)
        sheet = workbook[sheet_name]
    except KeyError:
        return 'sheet_not_found'

    schedule = {}

    for row_index, row in enumerate(sheet.iter_rows(values_only=True, min_row=2)):
        row_label = str(row[0])
        schedule[row_label] = {}

        # Match each column header (time) with value for that day
        for col_index, column in enumerate(sheet.iter_cols(values_only=True, min_col=2), start=1):
            header = str(column[0])
            value = column[row_index + 1]
            schedule[row_label][header] = value

    return schedule
